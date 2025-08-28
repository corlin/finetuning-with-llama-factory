"""
评估数据管理器

实现QA数据加载、验证、预处理和数据集准备功能。
支持训练QA数据格式，集成现有的数据模型和验证器。
"""

import json
import os
import re
from pathlib import Path
from typing import List, Dict, Any, Optional, Union, Tuple
from datetime import datetime
import logging
from dataclasses import asdict

from .interfaces import EvaluationDataManager as EvaluationDataManagerInterface
from .data_models import (
    QAEvaluationItem, 
    ValidationResult, 
    EvaluationDataset,
    ExpertEvaluationResult
)
from .config import ExpertiseLevel, EvaluationDimension
from .exceptions import DataFormatError, ValidationError
from ..data_models import TrainingExample, DifficultyLevel


class EvaluationDataManager(EvaluationDataManagerInterface):
    """评估数据管理器实现类"""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        初始化评估数据管理器
        
        Args:
            config: 配置参数
        """
        self.config = config or {}
        self.logger = logging.getLogger(__name__)
        
        # 支持的文件格式
        self.supported_formats = {'.json', '.jsonl', '.txt'}
        
        # 数据验证规则
        self.validation_rules = {
            'min_question_length': self.config.get('min_question_length', 5),
            'min_answer_length': self.config.get('min_answer_length', 10),
            'max_question_length': self.config.get('max_question_length', 2000),
            'max_answer_length': self.config.get('max_answer_length', 10000),
            'required_fields': ['instruction', 'output'],
            'optional_fields': ['input', 'thinking', 'system', 'history']
        }
    
    def load_qa_data(self, data_path: str) -> List[QAEvaluationItem]:
        """
        加载QA格式的测试数据
        
        Args:
            data_path: 数据文件路径
            
        Returns:
            List[QAEvaluationItem]: QA评估数据列表
            
        Raises:
            DataFormatError: 数据格式错误
            FileNotFoundError: 文件不存在
        """
        try:
            data_path = Path(data_path)
            
            if not data_path.exists():
                raise FileNotFoundError(f"数据文件不存在: {data_path}")
            
            if data_path.suffix not in self.supported_formats:
                raise DataFormatError(f"不支持的文件格式: {data_path.suffix}")
            
            # 根据文件格式加载数据
            if data_path.suffix == '.json':
                raw_data = self._load_json_data(data_path)
            elif data_path.suffix == '.jsonl':
                raw_data = self._load_jsonl_data(data_path)
            else:
                raise DataFormatError(f"暂不支持的格式: {data_path.suffix}")
            
            # 转换为QAEvaluationItem对象
            qa_items = self._convert_to_qa_items(raw_data, str(data_path))
            
            self.logger.info(f"成功加载 {len(qa_items)} 个QA评估项从 {data_path}")
            return qa_items
            
        except Exception as e:
            self.logger.error(f"加载QA数据失败: {e}")
            raise
    
    def _load_json_data(self, file_path: Path) -> List[Dict[str, Any]]:
        """加载JSON格式数据"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if isinstance(data, dict):
                return [data]
            elif isinstance(data, list):
                return data
            else:
                raise DataFormatError("JSON数据必须是字典或列表格式")
                
        except json.JSONDecodeError as e:
            raise DataFormatError(f"JSON格式错误: {e}")
    
    def _load_jsonl_data(self, file_path: Path) -> List[Dict[str, Any]]:
        """加载JSONL格式数据"""
        try:
            data = []
            with open(file_path, 'r', encoding='utf-8') as f:
                for line_num, line in enumerate(f, 1):
                    line = line.strip()
                    if line:
                        try:
                            item = json.loads(line)
                            data.append(item)
                        except json.JSONDecodeError as e:
                            self.logger.warning(f"跳过第{line_num}行，JSON格式错误: {e}")
            
            return data
            
        except Exception as e:
            raise DataFormatError(f"JSONL文件读取错误: {e}")
    
    def _convert_to_qa_items(self, raw_data: List[Dict[str, Any]], source_file: str) -> List[QAEvaluationItem]:
        """
        将原始数据转换为QAEvaluationItem对象
        
        Args:
            raw_data: 原始数据列表
            source_file: 源文件路径
            
        Returns:
            List[QAEvaluationItem]: 转换后的QA评估项列表
        """
        qa_items = []
        
        for idx, item in enumerate(raw_data):
            try:
                # 生成问题ID
                question_id = item.get('id') or f"{Path(source_file).stem}_{idx}"
                
                # 提取基本字段
                question = item.get('instruction', '').strip()
                context = item.get('input', '').strip() or None
                
                # 处理参考答案和模型答案
                reference_answer = item.get('output', '').strip()
                
                # 如果有thinking标签，提取最终答案作为参考答案
                if '<thinking>' in reference_answer and '</thinking>' in reference_answer:
                    # 提取thinking标签后的内容作为最终答案
                    thinking_end = reference_answer.rfind('</thinking>')
                    if thinking_end != -1:
                        final_answer = reference_answer[thinking_end + len('</thinking>'):].strip()
                        if final_answer:
                            reference_answer = final_answer
                
                # 模型答案初始为空，需要后续评估时填入
                model_answer = item.get('model_answer', '')
                
                # 提取领域标签
                domain_tags = self._extract_domain_tags(item)
                
                # 确定难度级别
                difficulty_level = self._determine_difficulty_level(item)
                
                # 提取期望概念
                expected_concepts = self._extract_expected_concepts(item)
                
                # 创建QAEvaluationItem
                qa_item = QAEvaluationItem(
                    question_id=question_id,
                    question=question,
                    context=context,
                    reference_answer=reference_answer,
                    model_answer=model_answer,
                    domain_tags=domain_tags,
                    difficulty_level=difficulty_level,
                    expected_concepts=expected_concepts,
                    metadata={
                        'source_file': source_file,
                        'original_index': idx,
                        'has_thinking': '<thinking>' in item.get('output', ''),
                        'system_prompt': item.get('system', ''),
                        'history': item.get('history', [])
                    }
                )
                
                qa_items.append(qa_item)
                
            except Exception as e:
                self.logger.warning(f"跳过第{idx}项数据，转换失败: {e}")
                continue
        
        return qa_items
    
    def _extract_domain_tags(self, item: Dict[str, Any]) -> List[str]:
        """提取领域标签"""
        tags = []
        
        # 从crypto_terms字段提取
        crypto_terms = item.get('crypto_terms', [])
        if crypto_terms:
            tags.extend(['密码学', '信息安全'])
        
        # 从内容中推断
        content = f"{item.get('instruction', '')} {item.get('output', '')}".lower()
        
        if any(keyword in content for keyword in ['密码', '加密', '签名', '哈希', '证书']):
            tags.append('密码学')
        
        if any(keyword in content for keyword in ['gb/t', '标准', '规范', '要求']):
            tags.append('标准规范')
        
        if any(keyword in content for keyword in ['系统', '网络', '应用', '数据']):
            tags.append('信息系统')
        
        return list(set(tags)) if tags else ['通用']
    
    def _determine_difficulty_level(self, item: Dict[str, Any]) -> ExpertiseLevel:
        """确定难度级别"""
        # 从原始数据的difficulty字段获取
        if 'difficulty' in item:
            difficulty_value = item['difficulty']
            if isinstance(difficulty_value, int):
                if difficulty_value == 1:
                    return ExpertiseLevel.BEGINNER
                elif difficulty_value == 2:
                    return ExpertiseLevel.INTERMEDIATE
                elif difficulty_value == 3:
                    return ExpertiseLevel.ADVANCED
                elif difficulty_value >= 4:
                    return ExpertiseLevel.EXPERT
        
        # 从difficulty_level字段获取
        if 'difficulty_level' in item:
            try:
                return ExpertiseLevel(item['difficulty_level'])
            except ValueError:
                pass
        
        # 根据内容长度和复杂度推断
        content = f"{item.get('instruction', '')} {item.get('output', '')}"
        
        if len(content) > 2000:
            return ExpertiseLevel.EXPERT
        elif len(content) > 1000:
            return ExpertiseLevel.ADVANCED
        elif len(content) > 500:
            return ExpertiseLevel.INTERMEDIATE
        else:
            return ExpertiseLevel.BEGINNER
    
    def _extract_expected_concepts(self, item: Dict[str, Any]) -> List[str]:
        """提取期望概念"""
        concepts = []
        
        # 从crypto_terms字段提取
        crypto_terms = item.get('crypto_terms', [])
        if crypto_terms:
            concepts.extend(crypto_terms)
        
        # 从内容中提取关键概念
        content = item.get('output', '')
        
        # 密码学相关概念
        crypto_concepts = [
            '机密性', '完整性', '真实性', '不可否认性',
            '对称加密', '非对称加密', '数字签名', '哈希函数',
            '密钥管理', '证书', 'PKI', 'CA'
        ]
        
        for concept in crypto_concepts:
            if concept in content:
                concepts.append(concept)
        
        return list(set(concepts))
    
    def validate_data_format(self, qa_data: List[QAEvaluationItem]) -> ValidationResult:
        """
        验证数据格式的正确性
        
        Args:
            qa_data: QA评估数据列表
            
        Returns:
            ValidationResult: 验证结果
        """
        validation_result = ValidationResult(is_valid=True)
        
        if not qa_data:
            validation_result.add_error("数据列表为空")
            return validation_result
        
        valid_count = 0
        invalid_count = 0
        
        for idx, item in enumerate(qa_data):
            item_errors = self._validate_single_item(item, idx)
            
            if item_errors:
                invalid_count += 1
                for error in item_errors:
                    validation_result.add_error(f"项目{idx}: {error}")
            else:
                valid_count += 1
        
        validation_result.valid_items_count = valid_count
        validation_result.invalid_items_count = invalid_count
        
        # 计算验证详情
        validation_result.validation_details = {
            'total_items': len(qa_data),
            'valid_percentage': (valid_count / len(qa_data)) * 100,
            'common_issues': self._analyze_common_issues(qa_data)
        }
        
        # 如果有效项目比例低于阈值，标记为无效
        valid_threshold = self.config.get('valid_threshold', 0.8)
        if (valid_count / len(qa_data)) < valid_threshold:
            validation_result.add_warning(f"有效项目比例 {valid_count/len(qa_data):.2%} 低于阈值 {valid_threshold:.2%}")
        
        self.logger.info(f"数据验证完成: {valid_count}/{len(qa_data)} 项有效")
        return validation_result
    
    def _validate_single_item(self, item: QAEvaluationItem, index: int) -> List[str]:
        """验证单个QA项目"""
        errors = []
        
        # 检查必需字段
        if not item.question_id.strip():
            errors.append("问题ID为空")
        
        if not item.question.strip():
            errors.append("问题内容为空")
        
        if not item.reference_answer.strip():
            errors.append("参考答案为空")
        
        # 检查长度限制
        if len(item.question) < self.validation_rules['min_question_length']:
            errors.append(f"问题长度过短: {len(item.question)} < {self.validation_rules['min_question_length']}")
        
        if len(item.question) > self.validation_rules['max_question_length']:
            errors.append(f"问题长度过长: {len(item.question)} > {self.validation_rules['max_question_length']}")
        
        if len(item.reference_answer) < self.validation_rules['min_answer_length']:
            errors.append(f"答案长度过短: {len(item.reference_answer)} < {self.validation_rules['min_answer_length']}")
        
        if len(item.reference_answer) > self.validation_rules['max_answer_length']:
            errors.append(f"答案长度过长: {len(item.reference_answer)} > {self.validation_rules['max_answer_length']}")
        
        # 检查数据质量
        if item.question == item.reference_answer:
            errors.append("问题和答案内容相同")
        
        # 检查特殊字符
        if self._contains_invalid_characters(item.question) or self._contains_invalid_characters(item.reference_answer):
            errors.append("包含无效字符")
        
        return errors
    
    def _contains_invalid_characters(self, text: str) -> bool:
        """检查是否包含无效字符"""
        # 检查控制字符（除了常见的换行符、制表符等）
        invalid_chars = ['\x00', '\x01', '\x02', '\x03', '\x04', '\x05', '\x06', '\x07', '\x08']
        return any(char in text for char in invalid_chars)
    
    def _analyze_common_issues(self, qa_data: List[QAEvaluationItem]) -> Dict[str, int]:
        """分析常见问题"""
        issues = {
            'empty_context': 0,
            'no_domain_tags': 0,
            'no_expected_concepts': 0,
            'short_questions': 0,
            'short_answers': 0
        }
        
        for item in qa_data:
            if not item.context:
                issues['empty_context'] += 1
            
            if not item.domain_tags:
                issues['no_domain_tags'] += 1
            
            if not item.expected_concepts:
                issues['no_expected_concepts'] += 1
            
            if len(item.question) < 20:
                issues['short_questions'] += 1
            
            if len(item.reference_answer) < 50:
                issues['short_answers'] += 1
        
        return issues   
 
    def prepare_evaluation_dataset(self, raw_data: List[Dict[str, Any]]) -> EvaluationDataset:
        """
        准备评估数据集
        
        Args:
            raw_data: 原始数据列表
            
        Returns:
            EvaluationDataset: 评估数据集
        """
        try:
            # 转换为QA评估项
            qa_items = self._convert_to_qa_items(raw_data, "prepared_dataset")
            
            # 生成数据集ID和名称
            dataset_id = f"eval_dataset_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            dataset_name = f"评估数据集 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # 创建数据集描述
            description = self._generate_dataset_description(qa_items)
            
            # 创建评估数据集
            dataset = EvaluationDataset(
                dataset_id=dataset_id,
                name=dataset_name,
                description=description,
                qa_items=qa_items
            )
            
            self.logger.info(f"成功准备评估数据集: {len(qa_items)} 个项目")
            return dataset
            
        except Exception as e:
            self.logger.error(f"准备评估数据集失败: {e}")
            raise ValidationError("dataset_preparation", [str(e)])
    
    def _generate_dataset_description(self, qa_items: List[QAEvaluationItem]) -> str:
        """生成数据集描述"""
        if not qa_items:
            return "空数据集"
        
        # 统计信息
        total_items = len(qa_items)
        
        # 难度分布
        difficulty_dist = {}
        for item in qa_items:
            level = item.difficulty_level.name
            difficulty_dist[level] = difficulty_dist.get(level, 0) + 1
        
        # 领域分布
        domain_dist = {}
        for item in qa_items:
            for tag in item.domain_tags:
                domain_dist[tag] = domain_dist.get(tag, 0) + 1
        
        # 生成描述
        description_parts = [
            f"包含 {total_items} 个QA评估项目",
            f"难度分布: {', '.join([f'{k}({v})' for k, v in difficulty_dist.items()])}",
            f"主要领域: {', '.join(list(domain_dist.keys())[:3])}"
        ]
        
        return "; ".join(description_parts)
    
    def export_results(self, results: ExpertEvaluationResult, format: str) -> str:
        """
        导出评估结果到不同格式
        
        Args:
            results: 评估结果
            format: 导出格式 (json, csv, xlsx)
            
        Returns:
            str: 导出文件路径
        """
        try:
            # 创建输出目录
            output_dir = Path(self.config.get('output_dir', './expert_evaluation_output'))
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # 生成文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"evaluation_result_{results.question_id}_{timestamp}"
            
            if format.lower() == 'json':
                return self._export_to_json(results, output_dir / f"{filename}.json")
            elif format.lower() == 'csv':
                return self._export_to_csv(results, output_dir / f"{filename}.csv")
            elif format.lower() == 'xlsx':
                return self._export_to_xlsx(results, output_dir / f"{filename}.xlsx")
            else:
                raise ValueError(f"不支持的导出格式: {format}")
                
        except Exception as e:
            self.logger.error(f"导出结果失败: {e}")
            raise
    
    def _export_to_json(self, results: ExpertEvaluationResult, file_path: Path) -> str:
        """导出为JSON格式"""
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(results.to_dict(), f, ensure_ascii=False, indent=2)
            
            self.logger.info(f"结果已导出到JSON文件: {file_path}")
            return str(file_path)
            
        except Exception as e:
            raise Exception(f"JSON导出失败: {e}")
    
    def _export_to_csv(self, results: ExpertEvaluationResult, file_path: Path) -> str:
        """导出为CSV格式"""
        try:
            import csv
            
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # 写入标题行
                writer.writerow(['指标', '分数', '置信度', '详情'])
                
                # 写入总体评分
                writer.writerow(['总体评分', results.overall_score, '', ''])
                
                # 写入各维度评分
                for dimension, score_obj in results.dimension_scores.items():
                    writer.writerow([
                        dimension.value,
                        score_obj.score,
                        score_obj.confidence,
                        str(score_obj.details)
                    ])
            
            self.logger.info(f"结果已导出到CSV文件: {file_path}")
            return str(file_path)
            
        except Exception as e:
            raise Exception(f"CSV导出失败: {e}")
    
    def _export_to_xlsx(self, results: ExpertEvaluationResult, file_path: Path) -> str:
        """导出为Excel格式"""
        try:
            # 尝试导入openpyxl，如果没有则降级到CSV
            try:
                import openpyxl
                from openpyxl import Workbook
            except ImportError:
                self.logger.warning("openpyxl未安装，降级到CSV格式")
                csv_path = file_path.with_suffix('.csv')
                return self._export_to_csv(results, csv_path)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "评估结果"
            
            # 设置标题行
            ws['A1'] = '指标'
            ws['B1'] = '分数'
            ws['C1'] = '置信度'
            ws['D1'] = '详情'
            
            # 写入数据
            row = 2
            ws[f'A{row}'] = '总体评分'
            ws[f'B{row}'] = results.overall_score
            
            row += 1
            for dimension, score_obj in results.dimension_scores.items():
                ws[f'A{row}'] = dimension.value
                ws[f'B{row}'] = score_obj.score
                ws[f'C{row}'] = score_obj.confidence
                ws[f'D{row}'] = str(score_obj.details)
                row += 1
            
            wb.save(file_path)
            
            self.logger.info(f"结果已导出到Excel文件: {file_path}")
            return str(file_path)
            
        except Exception as e:
            raise Exception(f"Excel导出失败: {e}")
    
    def batch_load_datasets(self, data_paths: List[str]) -> List[EvaluationDataset]:
        """
        批量加载多个数据集
        
        Args:
            data_paths: 数据文件路径列表
            
        Returns:
            List[EvaluationDataset]: 评估数据集列表
        """
        datasets = []
        
        for path in data_paths:
            try:
                qa_items = self.load_qa_data(path)
                
                # 创建数据集
                dataset_id = f"dataset_{Path(path).stem}"
                dataset_name = f"数据集 - {Path(path).name}"
                description = f"从文件 {path} 加载的数据集"
                
                dataset = EvaluationDataset(
                    dataset_id=dataset_id,
                    name=dataset_name,
                    description=description,
                    qa_items=qa_items
                )
                
                datasets.append(dataset)
                
            except Exception as e:
                self.logger.error(f"加载数据集失败 {path}: {e}")
                continue
        
        self.logger.info(f"批量加载完成: {len(datasets)}/{len(data_paths)} 个数据集成功")
        return datasets
    
    def create_balanced_dataset(self, 
                              qa_items: List[QAEvaluationItem],
                              target_size: int = 100,
                              balance_by: str = 'difficulty') -> EvaluationDataset:
        """
        创建平衡的评估数据集
        
        Args:
            qa_items: 原始QA项目列表
            target_size: 目标数据集大小
            balance_by: 平衡依据 ('difficulty', 'domain', 'random')
            
        Returns:
            EvaluationDataset: 平衡的评估数据集
        """
        try:
            if balance_by == 'difficulty':
                balanced_items = self._balance_by_difficulty(qa_items, target_size)
            elif balance_by == 'domain':
                balanced_items = self._balance_by_domain(qa_items, target_size)
            elif balance_by == 'random':
                balanced_items = self._random_sample(qa_items, target_size)
            else:
                raise ValueError(f"不支持的平衡方式: {balance_by}")
            
            # 创建平衡数据集
            dataset_id = f"balanced_dataset_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            dataset_name = f"平衡数据集 ({balance_by})"
            description = f"基于{balance_by}平衡的评估数据集，包含{len(balanced_items)}个项目"
            
            dataset = EvaluationDataset(
                dataset_id=dataset_id,
                name=dataset_name,
                description=description,
                qa_items=balanced_items
            )
            
            self.logger.info(f"创建平衡数据集成功: {len(balanced_items)} 个项目")
            return dataset
            
        except Exception as e:
            self.logger.error(f"创建平衡数据集失败: {e}")
            raise
    
    def _balance_by_difficulty(self, qa_items: List[QAEvaluationItem], target_size: int) -> List[QAEvaluationItem]:
        """按难度平衡数据集"""
        import random
        
        # 按难度分组
        difficulty_groups = {}
        for item in qa_items:
            level = item.difficulty_level
            if level not in difficulty_groups:
                difficulty_groups[level] = []
            difficulty_groups[level].append(item)
        
        # 计算每个难度级别的目标数量
        num_levels = len(difficulty_groups)
        items_per_level = target_size // num_levels
        remainder = target_size % num_levels
        
        balanced_items = []
        for i, (level, items) in enumerate(difficulty_groups.items()):
            # 为前几个级别分配余数
            current_target = items_per_level + (1 if i < remainder else 0)
            
            if len(items) >= current_target:
                selected = random.sample(items, current_target)
            else:
                selected = items  # 如果不够就全部选择
            
            balanced_items.extend(selected)
        
        return balanced_items
    
    def _balance_by_domain(self, qa_items: List[QAEvaluationItem], target_size: int) -> List[QAEvaluationItem]:
        """按领域平衡数据集"""
        import random
        
        # 按主要领域分组（取第一个标签）
        domain_groups = {}
        for item in qa_items:
            domain = item.domain_tags[0] if item.domain_tags else '未分类'
            if domain not in domain_groups:
                domain_groups[domain] = []
            domain_groups[domain].append(item)
        
        # 计算每个领域的目标数量
        num_domains = len(domain_groups)
        items_per_domain = target_size // num_domains
        remainder = target_size % num_domains
        
        balanced_items = []
        for i, (domain, items) in enumerate(domain_groups.items()):
            current_target = items_per_domain + (1 if i < remainder else 0)
            
            if len(items) >= current_target:
                selected = random.sample(items, current_target)
            else:
                selected = items
            
            balanced_items.extend(selected)
        
        return balanced_items
    
    def _random_sample(self, qa_items: List[QAEvaluationItem], target_size: int) -> List[QAEvaluationItem]:
        """随机采样"""
        import random
        
        if len(qa_items) <= target_size:
            return qa_items
        
        return random.sample(qa_items, target_size)
    
    def get_dataset_statistics(self, dataset: EvaluationDataset) -> Dict[str, Any]:
        """
        获取数据集统计信息
        
        Args:
            dataset: 评估数据集
            
        Returns:
            Dict[str, Any]: 统计信息
        """
        if not dataset.qa_items:
            return {'total_items': 0}
        
        stats = {
            'total_items': len(dataset.qa_items),
            'difficulty_distribution': {},
            'domain_distribution': {},
            'question_length_stats': {},
            'answer_length_stats': {},
            'concept_coverage': {}
        }
        
        # 难度分布
        for item in dataset.qa_items:
            level = item.difficulty_level.name
            stats['difficulty_distribution'][level] = stats['difficulty_distribution'].get(level, 0) + 1
        
        # 领域分布
        for item in dataset.qa_items:
            for tag in item.domain_tags:
                stats['domain_distribution'][tag] = stats['domain_distribution'].get(tag, 0) + 1
        
        # 长度统计
        question_lengths = [len(item.question) for item in dataset.qa_items]
        answer_lengths = [len(item.reference_answer) for item in dataset.qa_items]
        
        stats['question_length_stats'] = {
            'min': min(question_lengths),
            'max': max(question_lengths),
            'avg': sum(question_lengths) / len(question_lengths),
            'median': sorted(question_lengths)[len(question_lengths) // 2]
        }
        
        stats['answer_length_stats'] = {
            'min': min(answer_lengths),
            'max': max(answer_lengths),
            'avg': sum(answer_lengths) / len(answer_lengths),
            'median': sorted(answer_lengths)[len(answer_lengths) // 2]
        }
        
        # 概念覆盖
        all_concepts = []
        for item in dataset.qa_items:
            all_concepts.extend(item.expected_concepts)
        
        concept_counts = {}
        for concept in all_concepts:
            concept_counts[concept] = concept_counts.get(concept, 0) + 1
        
        stats['concept_coverage'] = dict(sorted(concept_counts.items(), key=lambda x: x[1], reverse=True)[:10])
        
        return stats


# 测试和验证函数
def test_dataset_preparation():
    """测试数据集准备功能"""
    try:
        # 创建测试数据管理器
        manager = EvaluationDataManager()
        
        # 测试数据
        test_data = [
            {
                "instruction": "什么是对称加密？",
                "input": "",
                "output": "对称加密是使用相同密钥进行加密和解密的加密方式。",
                "model_answer": "对称加密是一种加密算法类型。",
                "crypto_terms": ["对称加密", "密钥"],
                "difficulty": 2
            },
            {
                "instruction": "解释数字签名的工作原理",
                "input": "",
                "output": "<thinking>\n数字签名涉及非对称加密技术...\n</thinking>\n数字签名使用私钥签名，公钥验证。",
                "model_answer": "数字签名通过非对称加密实现身份验证。",
                "crypto_terms": ["数字签名", "非对称加密"],
                "difficulty": 3
            }
        ]
        
        # 测试数据集准备
        dataset = manager.prepare_evaluation_dataset(test_data)
        print(f"✓ 数据集准备成功: {len(dataset.qa_items)} 个项目")
        
        # 测试数据验证
        validation_result = manager.validate_data_format(dataset.qa_items)
        print(f"✓ 数据验证完成: {validation_result.valid_items_count}/{validation_result.valid_items_count + validation_result.invalid_items_count} 项有效")
        
        # 测试统计信息
        if dataset.qa_items:
            stats = manager.get_dataset_statistics(dataset)
            print(f"✓ 统计信息: {stats['total_items']} 个项目，难度分布: {stats['difficulty_distribution']}")
        else:
            print("⚠ 数据集为空，跳过统计信息测试")
        
        print("✓ 数据集准备功能测试通过")
        return True
        
    except Exception as e:
        print(f"✗ 数据集准备功能测试失败: {e}")
        return False


# 高级数据集准备功能
class AdvancedDatasetPreparator:
    """高级数据集准备器"""
    
    def __init__(self, data_manager: EvaluationDataManager):
        """
        初始化高级数据集准备器
        
        Args:
            data_manager: 基础数据管理器
        """
        self.data_manager = data_manager
        self.logger = logging.getLogger(__name__)
    
    def prepare_stratified_dataset(self, 
                                 qa_items: List[QAEvaluationItem],
                                 stratify_by: str = 'difficulty',
                                 train_ratio: float = 0.7,
                                 val_ratio: float = 0.15,
                                 test_ratio: float = 0.15) -> Tuple[EvaluationDataset, EvaluationDataset, EvaluationDataset]:
        """
        准备分层数据集
        
        Args:
            qa_items: QA项目列表
            stratify_by: 分层依据 ('difficulty', 'domain', 'length')
            train_ratio: 训练集比例
            val_ratio: 验证集比例
            test_ratio: 测试集比例
            
        Returns:
            Tuple[EvaluationDataset, EvaluationDataset, EvaluationDataset]: 训练集、验证集、测试集
        """
        try:
            if abs(train_ratio + val_ratio + test_ratio - 1.0) > 0.01:
                raise ValueError("分割比例总和必须为1.0")
            
            # 按指定标准分组
            if stratify_by == 'difficulty':
                groups = self._group_by_difficulty(qa_items)
            elif stratify_by == 'domain':
                groups = self._group_by_domain(qa_items)
            elif stratify_by == 'length':
                groups = self._group_by_length(qa_items)
            else:
                raise ValueError(f"不支持的分层标准: {stratify_by}")
            
            # 分层采样
            train_items, val_items, test_items = [], [], []
            
            for group_name, group_items in groups.items():
                # 计算每组的分割数量
                total = len(group_items)
                train_size = int(total * train_ratio)
                val_size = int(total * val_ratio)
                
                # 随机打乱
                import random
                shuffled = group_items.copy()
                random.shuffle(shuffled)
                
                # 分割
                train_items.extend(shuffled[:train_size])
                val_items.extend(shuffled[train_size:train_size + val_size])
                test_items.extend(shuffled[train_size + val_size:])
            
            # 创建数据集
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            train_dataset = EvaluationDataset(
                dataset_id=f"train_{stratify_by}_{timestamp}",
                name=f"训练集 (分层: {stratify_by})",
                description=f"基于{stratify_by}分层的训练数据集，包含{len(train_items)}个项目",
                qa_items=train_items
            )
            
            val_dataset = EvaluationDataset(
                dataset_id=f"val_{stratify_by}_{timestamp}",
                name=f"验证集 (分层: {stratify_by})",
                description=f"基于{stratify_by}分层的验证数据集，包含{len(val_items)}个项目",
                qa_items=val_items
            )
            
            test_dataset = EvaluationDataset(
                dataset_id=f"test_{stratify_by}_{timestamp}",
                name=f"测试集 (分层: {stratify_by})",
                description=f"基于{stratify_by}分层的测试数据集，包含{len(test_items)}个项目",
                qa_items=test_items
            )
            
            self.logger.info(f"分层数据集准备完成: 训练{len(train_items)}, 验证{len(val_items)}, 测试{len(test_items)}")
            return train_dataset, val_dataset, test_dataset
            
        except Exception as e:
            self.logger.error(f"分层数据集准备失败: {e}")
            raise
    
    def _group_by_difficulty(self, qa_items: List[QAEvaluationItem]) -> Dict[str, List[QAEvaluationItem]]:
        """按难度分组"""
        groups = {}
        for item in qa_items:
            level = item.difficulty_level.name
            if level not in groups:
                groups[level] = []
            groups[level].append(item)
        return groups
    
    def _group_by_domain(self, qa_items: List[QAEvaluationItem]) -> Dict[str, List[QAEvaluationItem]]:
        """按领域分组"""
        groups = {}
        for item in qa_items:
            # 使用第一个领域标签作为主要分组
            domain = item.domain_tags[0] if item.domain_tags else '未分类'
            if domain not in groups:
                groups[domain] = []
            groups[domain].append(item)
        return groups
    
    def _group_by_length(self, qa_items: List[QAEvaluationItem]) -> Dict[str, List[QAEvaluationItem]]:
        """按长度分组"""
        groups = {'短': [], '中': [], '长': []}
        
        for item in qa_items:
            total_length = len(item.question) + len(item.reference_answer)
            if total_length < 200:
                groups['短'].append(item)
            elif total_length < 800:
                groups['中'].append(item)
            else:
                groups['长'].append(item)
        
        return groups
    
    def create_cross_validation_datasets(self, 
                                       qa_items: List[QAEvaluationItem],
                                       k_folds: int = 5) -> List[Tuple[EvaluationDataset, EvaluationDataset]]:
        """
        创建K折交叉验证数据集
        
        Args:
            qa_items: QA项目列表
            k_folds: 折数
            
        Returns:
            List[Tuple[EvaluationDataset, EvaluationDataset]]: (训练集, 验证集) 对列表
        """
        try:
            import random
            
            # 随机打乱数据
            shuffled_items = qa_items.copy()
            random.shuffle(shuffled_items)
            
            # 计算每折的大小
            fold_size = len(shuffled_items) // k_folds
            
            cv_datasets = []
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            for fold in range(k_folds):
                # 确定验证集范围
                val_start = fold * fold_size
                val_end = val_start + fold_size if fold < k_folds - 1 else len(shuffled_items)
                
                # 分割数据
                val_items = shuffled_items[val_start:val_end]
                train_items = shuffled_items[:val_start] + shuffled_items[val_end:]
                
                # 创建数据集
                train_dataset = EvaluationDataset(
                    dataset_id=f"cv_train_fold{fold}_{timestamp}",
                    name=f"交叉验证训练集 - 第{fold+1}折",
                    description=f"K折交叉验证训练数据集 (K={k_folds}, 折={fold+1})",
                    qa_items=train_items
                )
                
                val_dataset = EvaluationDataset(
                    dataset_id=f"cv_val_fold{fold}_{timestamp}",
                    name=f"交叉验证验证集 - 第{fold+1}折",
                    description=f"K折交叉验证验证数据集 (K={k_folds}, 折={fold+1})",
                    qa_items=val_items
                )
                
                cv_datasets.append((train_dataset, val_dataset))
            
            self.logger.info(f"K折交叉验证数据集创建完成: {k_folds} 折")
            return cv_datasets
            
        except Exception as e:
            self.logger.error(f"交叉验证数据集创建失败: {e}")
            raise
    
    def augment_dataset(self, 
                       qa_items: List[QAEvaluationItem],
                       augmentation_methods: List[str] = None) -> List[QAEvaluationItem]:
        """
        数据增强
        
        Args:
            qa_items: 原始QA项目列表
            augmentation_methods: 增强方法列表
            
        Returns:
            List[QAEvaluationItem]: 增强后的QA项目列表
        """
        if augmentation_methods is None:
            augmentation_methods = ['paraphrase', 'context_variation']
        
        augmented_items = qa_items.copy()
        
        for method in augmentation_methods:
            if method == 'paraphrase':
                augmented_items.extend(self._paraphrase_questions(qa_items))
            elif method == 'context_variation':
                augmented_items.extend(self._vary_context(qa_items))
            elif method == 'difficulty_scaling':
                augmented_items.extend(self._scale_difficulty(qa_items))
        
        self.logger.info(f"数据增强完成: {len(qa_items)} -> {len(augmented_items)} 个项目")
        return augmented_items
    
    def _paraphrase_questions(self, qa_items: List[QAEvaluationItem]) -> List[QAEvaluationItem]:
        """问题改写增强"""
        paraphrased = []
        
        paraphrase_patterns = [
            ("什么是", "请解释"),
            ("如何", "怎样"),
            ("为什么", "什么原因"),
            ("描述", "说明"),
            ("解释", "阐述")
        ]
        
        for item in qa_items:
            for old_pattern, new_pattern in paraphrase_patterns:
                if old_pattern in item.question:
                    new_question = item.question.replace(old_pattern, new_pattern, 1)
                    
                    paraphrased_item = QAEvaluationItem(
                        question_id=f"{item.question_id}_para",
                        question=new_question,
                        context=item.context,
                        reference_answer=item.reference_answer,
                        model_answer=item.model_answer,
                        domain_tags=item.domain_tags,
                        difficulty_level=item.difficulty_level,
                        expected_concepts=item.expected_concepts,
                        metadata={**item.metadata, 'augmentation': 'paraphrase'}
                    )
                    paraphrased.append(paraphrased_item)
                    break  # 只应用第一个匹配的模式
        
        return paraphrased
    
    def _vary_context(self, qa_items: List[QAEvaluationItem]) -> List[QAEvaluationItem]:
        """上下文变化增强"""
        varied = []
        
        context_variations = [
            "在企业环境中，",
            "从技术角度来看，",
            "在实际应用中，",
            "考虑到安全性，"
        ]
        
        for item in qa_items:
            if not item.context:  # 只对没有上下文的项目添加
                import random
                new_context = random.choice(context_variations)
                
                varied_item = QAEvaluationItem(
                    question_id=f"{item.question_id}_ctx",
                    question=item.question,
                    context=new_context,
                    reference_answer=item.reference_answer,
                    model_answer=item.model_answer,
                    domain_tags=item.domain_tags,
                    difficulty_level=item.difficulty_level,
                    expected_concepts=item.expected_concepts,
                    metadata={**item.metadata, 'augmentation': 'context_variation'}
                )
                varied.append(varied_item)
        
        return varied
    
    def _scale_difficulty(self, qa_items: List[QAEvaluationItem]) -> List[QAEvaluationItem]:
        """难度调整增强"""
        scaled = []
        
        for item in qa_items:
            if item.difficulty_level == ExpertiseLevel.BEGINNER:
                # 提升到中级
                scaled_item = QAEvaluationItem(
                    question_id=f"{item.question_id}_up",
                    question=f"深入分析：{item.question}",
                    context=item.context,
                    reference_answer=item.reference_answer,
                    model_answer=item.model_answer,
                    domain_tags=item.domain_tags,
                    difficulty_level=ExpertiseLevel.INTERMEDIATE,
                    expected_concepts=item.expected_concepts,
                    metadata={**item.metadata, 'augmentation': 'difficulty_up'}
                )
                scaled.append(scaled_item)
            
            elif item.difficulty_level == ExpertiseLevel.ADVANCED:
                # 降低到中级
                scaled_item = QAEvaluationItem(
                    question_id=f"{item.question_id}_down",
                    question=f"简要说明：{item.question}",
                    context=item.context,
                    reference_answer=item.reference_answer,
                    model_answer=item.model_answer,
                    domain_tags=item.domain_tags,
                    difficulty_level=ExpertiseLevel.INTERMEDIATE,
                    expected_concepts=item.expected_concepts,
                    metadata={**item.metadata, 'augmentation': 'difficulty_down'}
                )
                scaled.append(scaled_item)
        
        return scaled
    
    def create_benchmark_dataset(self, 
                               qa_items: List[QAEvaluationItem],
                               benchmark_type: str = 'comprehensive') -> EvaluationDataset:
        """
        创建基准测试数据集
        
        Args:
            qa_items: QA项目列表
            benchmark_type: 基准类型 ('comprehensive', 'domain_specific', 'difficulty_ladder')
            
        Returns:
            EvaluationDataset: 基准测试数据集
        """
        try:
            if benchmark_type == 'comprehensive':
                selected_items = self._select_comprehensive_benchmark(qa_items)
            elif benchmark_type == 'domain_specific':
                selected_items = self._select_domain_benchmark(qa_items)
            elif benchmark_type == 'difficulty_ladder':
                selected_items = self._select_difficulty_ladder(qa_items)
            else:
                raise ValueError(f"不支持的基准类型: {benchmark_type}")
            
            # 创建基准数据集
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            benchmark_dataset = EvaluationDataset(
                dataset_id=f"benchmark_{benchmark_type}_{timestamp}",
                name=f"基准测试数据集 ({benchmark_type})",
                description=f"{benchmark_type}类型的基准测试数据集，包含{len(selected_items)}个精选项目",
                qa_items=selected_items
            )
            
            self.logger.info(f"基准数据集创建完成: {benchmark_type} 类型，{len(selected_items)} 个项目")
            return benchmark_dataset
            
        except Exception as e:
            self.logger.error(f"基准数据集创建失败: {e}")
            raise
    
    def _select_comprehensive_benchmark(self, qa_items: List[QAEvaluationItem]) -> List[QAEvaluationItem]:
        """选择综合基准项目"""
        # 确保每个难度级别和领域都有代表
        selected = []
        
        # 按难度和领域分组
        groups = {}
        for item in qa_items:
            difficulty = item.difficulty_level.name
            domain = item.domain_tags[0] if item.domain_tags else '通用'
            key = f"{difficulty}_{domain}"
            
            if key not in groups:
                groups[key] = []
            groups[key].append(item)
        
        # 从每组选择代表性项目
        import random
        for group_items in groups.values():
            if group_items:
                # 选择最多2个代表性项目
                num_select = min(2, len(group_items))
                selected.extend(random.sample(group_items, num_select))
        
        return selected
    
    def _select_domain_benchmark(self, qa_items: List[QAEvaluationItem]) -> List[QAEvaluationItem]:
        """选择领域特定基准项目"""
        # 按领域分组，每个领域选择不同难度的项目
        domain_groups = {}
        for item in qa_items:
            domain = item.domain_tags[0] if item.domain_tags else '通用'
            if domain not in domain_groups:
                domain_groups[domain] = []
            domain_groups[domain].append(item)
        
        selected = []
        import random
        
        for domain, items in domain_groups.items():
            # 按难度分组
            difficulty_groups = {}
            for item in items:
                diff = item.difficulty_level.name
                if diff not in difficulty_groups:
                    difficulty_groups[diff] = []
                difficulty_groups[diff].append(item)
            
            # 每个难度选择1个项目
            for diff_items in difficulty_groups.values():
                if diff_items:
                    selected.append(random.choice(diff_items))
        
        return selected
    
    def _select_difficulty_ladder(self, qa_items: List[QAEvaluationItem]) -> List[QAEvaluationItem]:
        """选择难度阶梯基准项目"""
        # 按难度排序，选择渐进式难度的项目
        difficulty_order = [
            ExpertiseLevel.BEGINNER,
            ExpertiseLevel.INTERMEDIATE,
            ExpertiseLevel.ADVANCED,
            ExpertiseLevel.EXPERT
        ]
        
        selected = []
        import random
        
        for difficulty in difficulty_order:
            matching_items = [item for item in qa_items if item.difficulty_level == difficulty]
            if matching_items:
                # 每个难度级别选择3个项目
                num_select = min(3, len(matching_items))
                selected.extend(random.sample(matching_items, num_select))
        
        return selected


def test_advanced_dataset_preparation():
    """测试高级数据集准备功能"""
    try:
        # 创建基础数据管理器
        base_manager = EvaluationDataManager()
        
        # 创建高级准备器
        preparator = AdvancedDatasetPreparator(base_manager)
        
        # 创建测试数据
        test_data = []
        for i in range(10):
            test_data.append({
                "instruction": f"测试问题 {i+1}",
                "input": "",
                "output": f"测试答案 {i+1}",
                "model_answer": f"模型答案 {i+1}",
                "crypto_terms": ["测试术语"],
                "difficulty": (i % 4) + 1  # 1-4的难度
            })
        
        # 准备基础数据集
        dataset = base_manager.prepare_evaluation_dataset(test_data)
        print(f"✓ 基础数据集准备: {len(dataset.qa_items)} 个项目")
        
        # 测试分层数据集
        train_ds, val_ds, test_ds = preparator.prepare_stratified_dataset(
            dataset.qa_items, stratify_by='difficulty'
        )
        print(f"✓ 分层数据集: 训练{len(train_ds.qa_items)}, 验证{len(val_ds.qa_items)}, 测试{len(test_ds.qa_items)}")
        
        # 测试交叉验证数据集
        cv_datasets = preparator.create_cross_validation_datasets(dataset.qa_items, k_folds=3)
        print(f"✓ 交叉验证数据集: {len(cv_datasets)} 折")
        
        # 测试数据增强
        augmented_items = preparator.augment_dataset(dataset.qa_items[:3])
        print(f"✓ 数据增强: {len(dataset.qa_items[:3])} -> {len(augmented_items)} 个项目")
        
        # 测试基准数据集
        benchmark_ds = preparator.create_benchmark_dataset(dataset.qa_items, 'comprehensive')
        print(f"✓ 基准数据集: {len(benchmark_ds.qa_items)} 个项目")
        
        print("✓ 高级数据集准备功能测试通过")
        return True
        
    except Exception as e:
        print(f"✗ 高级数据集准备功能测试失败: {e}")
        return False


# 导出主要测试函数
def test_dataset_preparation():
    """完整的数据集准备功能测试"""
    print("=== 基础数据集准备测试 ===")
    basic_result = test_dataset_preparation_basic()
    
    print("\n=== 高级数据集准备测试 ===")
    advanced_result = test_advanced_dataset_preparation()
    
    return basic_result and advanced_result


def test_dataset_preparation_basic():
    """基础数据集准备功能测试"""
    try:
        # 创建测试数据管理器
        manager = EvaluationDataManager()
        
        # 测试数据
        test_data = [
            {
                "instruction": "什么是对称加密？",
                "input": "",
                "output": "对称加密是使用相同密钥进行加密和解密的加密方式。",
                "model_answer": "对称加密是一种加密算法类型。",
                "crypto_terms": ["对称加密", "密钥"],
                "difficulty": 2
            },
            {
                "instruction": "解释数字签名的工作原理",
                "input": "",
                "output": "<thinking>\n数字签名涉及非对称加密技术...\n</thinking>\n数字签名使用私钥签名，公钥验证。",
                "model_answer": "数字签名通过非对称加密实现身份验证。",
                "crypto_terms": ["数字签名", "非对称加密"],
                "difficulty": 3
            }
        ]
        
        # 测试数据集准备
        dataset = manager.prepare_evaluation_dataset(test_data)
        print(f"✓ 数据集准备成功: {len(dataset.qa_items)} 个项目")
        
        # 测试数据验证
        validation_result = manager.validate_data_format(dataset.qa_items)
        print(f"✓ 数据验证完成: {validation_result.valid_items_count}/{validation_result.valid_items_count + validation_result.invalid_items_count} 项有效")
        
        # 测试统计信息
        if dataset.qa_items:
            stats = manager.get_dataset_statistics(dataset)
            print(f"✓ 统计信息: {stats['total_items']} 个项目，难度分布: {stats['difficulty_distribution']}")
        else:
            print("⚠ 数据集为空，跳过统计信息测试")
        
        print("✓ 基础数据集准备功能测试通过")
        return True
        
    except Exception as e:
        print(f"✗ 基础数据集准备功能测试失败: {e}")
        return False

if __name__ == "__main__":
    # 运行测试
    test_dataset_preparation()